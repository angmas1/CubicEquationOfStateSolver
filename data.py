import numpy
import matplotlib.pyplot as plt
import pandas
import random
import os
import sys
from sympy import *
import openpyxl


'''
print("ho");
print("\n"*5)
print("%s\n%s\n%s" %("hey","how","are"))
print("a","b","c")

gl=["hello","hi",'5']
print(gl[2])
print(gl[0:2])

ol=["bhak","bhenchod","chutiye"]
gl.append("fuck")q
print(gl)
biglist=[ol,gl]
biglist2=ol+gl
print(biglist)
print(biglist2)
gl.insert(2,"you")
gl.append(3)
gl.remove(3)
gl.sort()
print(gl)
gl.reverse()
print(gl)
print(biglist)
print(biglist2)
print(len(biglist))
print(len(biglist2))
print(max(biglist2))
print(min(biglist2))


tuple1=(3,1,4,1,5,9)
print(tuple1)
list=list(tuple1)
print(list)
atup=tuple(list)
print(atup)

hero = {"iron man":"tony stark",
        "captain america":"steve rogers",
        "black widow":"natasha romanoff",
        "hulk":"bruce banner"}
hero["iron man"]="tony"
print(hero["iron man"])
del hero["hulk"]

print(hero.values())



age=60

if age<16:
    print(age)
elif age>40:
    print("too old")
else:
    print("neub")


if (True and not False):
    print("hello")

for x in range(0,10):
    print(x,end="")

list=[1,2,3,4,5,6,7,8,9]
print("\n\n")

for x in range(0,len(list)):
    print(x,list[x])

rownum=10

for x in range(0,rownum):
    for y in range(0,rownum-1-x):
        print(" ",end="")
    for z in range(0,2*x+1):
        print("*",end="")
    print("\n",end="")

rand=random.randrange(0,1000)

while(rand!=15):
    print(rand)
    rand=random.randrange(0,1000)
print(rand)


def addnum(a,b):
    s=a+b
    return s

x=5
y=6

print(addnum(x,y))
print(x,y)



print("What is your name?")

name=sys.stdin.readline()

print("Hello",name)



string="Hello how are you I am fine thank you"
print(string[0:4])
print(string[:8])
print(string[-5:])
x= sympy.symbols('x')
print(sympy.diff(x**3,x))



print("%c is my %s letter And %d is number and %.5f is float" %('a',"character",5,5.234532423))

string="    hello how are you I am fine thank you    "

print(string.capitalize())
print(string.find("are "))
print(string.isalpha())
print(string.isalnum())
print(string.replace("hello ","  hi  "))
print(string.strip())

ls=string.split(" ")
print(ls)




writing=open("test.txt","a")

a=writing.write("Hello\n")
print(a)
writing.close()

reading=open("test.txt","r")#ab+ readwrite wb write r+ read

print(reading.mode)
print(reading.name)

#testfile.write(bytes("\nAdding this to the file.\n","UTF-8"))
a=reading.read()
print(a)
reading.close()

os.remove("test.txt")


class Animal:
    __name=""
    __height=0
    __weight=0
    __sound=0

    def __init__(self,name,height,weight,sound):
        self.__name=name
        self.__height=height
        self.__weight=weight
        self.__sound=sound

    def set_name(self, name):
        self.__name = name

    def set_height(self, height):
        self.__height = height

    def set_weight(self, height):
        self.__height = height

    def set_sound(self, sound):
        self.__sound = sound

    def get_name(self):
        return self.__name

    def get_height(self):
        return str(self.__height)

    def get_weight(self):
        return str(self.__weight)

    def get_sound(self):
        return self.__sound
    def returnall(self):
        return "{} {} {} {}".format(self.__name,self.__height,self.__weight,self.__sound)


cat = Animal("rux",33,10,"meow")

print(cat.returnall())

class Dog(Animal):
    __owner = ""

    def __init__(self,a,b,c,d,e):
        self.__owner=e
        super(Dog,self).__init__(a,b,c,d)

    def returnall(self):
        return "{} {} {} {} {}".format(self.get_name(),self.get_height(),self.get_weight(),self.get_sound(),self.__owner)

    def multiple_sounds(self, how_many=None):
        if how_many == None:
            print(self.get_sound)
        else:
            print(self.get_sound() * how_many)

d1=Dog("hey",1,2,"wow","me")
print(d1.returnall())


class Animal:
    # None signifies the lack of a value
    # You can make a variable private by starting it with __
    __name = None
    __height = None
    __weight = None
    __sound = None

    # The constructor is called to set up or initialize an object
    # self allows an object to refer to itself inside of the class
    def __init__(self, name, height, weight, sound):
        self.__name = name
        self.__height = height
        self.__weight = weight
        self.__sound = sound

    def set_name(self, name):
        self.__name = name

    def set_height(self, height):
        self.__height = height

    def set_weight(self, height):
        self.__height = height

    def set_sound(self, sound):
        self.__sound = sound

    def get_name(self):
        return self.__name

    def get_height(self):
        return str(self.__height)

    def get_weight(self):
        return str(self.__weight)

    def get_sound(self):
        return self.__sound

    def get_type(self):
        print("Animal")

    def toString(self):
        return "{} is {} cm tall and {} kilograms and says {}".format(self.__name, self.__height, self.__weight,
                                                                      self.__sound)


# How to create a Animal object
cat = Animal('Whiskers', 33, 10, 'Meow')

print(cat.toString())


# You can't access this value directly because it is private
# print(cat.__name)

# INHERITANCE -------------
# You can inherit all of the variables and methods from another class

class Dog(Animal):
    __owner = None

    def __init__(self, name, height, weight, sound, owner):
        self.__owner = owner

        # How to call the super class constructor
        super(Dog, self).__init__(name, height, weight, sound)

    def set_owner(self, owner):
        self.__owner = owner

    def get_owner(self):
        return self.__owner

    def get_type(self):
        print("Dog")

    # We can overwrite functions in the super class
    def toString(self):
        return "{} is {} cm tall and {} kilograms and says {}. His owner is {}".format(self.get_name(),
                                                                                       self.get_height(),
                                                                                       self.get_weight(),
                                                                                       self.get_sound(), 
                                                                                       self.__owner)

    # You don't have to require attributes to be sent
    # This allows for method overloading
    def multiple_sounds(self, how_many=None):
        if how_many == None:
            print(self.get_sound)
        else:
            print(self.get_sound() * how_many)


spot = Dog("Spot", 53, 27, "Ruff", "Derek")

print(spot.toString())
print(spot.multiple_sounds(5))


x=[1,2,3,4,5,6,7,8,9,10]
y=[3,5,2,6,3,8,2,6,5,1]

plt.bar(x,y)
plt.show()



rownum=10

for x in range(0,rownum):
    for y in range(0,rownum-1-x):
        print(" ",end="")
    for z in range(0,2*x+1):
        print("*",end="")
    print("\n",end="")
    
    
    


wb = load_workbook("data.xlsx")
sheet=wb['Sheet1']
column = sheet['A']
print(column[25].value)
#sheet_ranges = wb['A1']

wb = openpyxl.Workbook()

sheet = wb.active

c1 = sheet.cell(row=1, column=1)

c1.value = "ANKIT"

c2 = sheet.cell(row=1, column=2)
c2.value = "RAI"

c3 = sheet['A2']
c3.value = "RAHUL"

c4 = sheet['B2']
c4.value = "RAI"

wb.save("C:\\Users\\Anunay\\Desktop\\demo.xlsx")'''

def function(v,p):
    a = 0.000000311
    pid = 70
    sh = 0.04 * (2300 ** 0.75) * (0.7 ** 0.33)
    pif = p
    d = 1.022 * (10 ** -9)
    k = sh * d / 0.026
    K = 2.24 * (10 ** 5)
    return (a*pid*exp(-v*(10**-6)/k)-a*pif*exp(v*(10**-6)*K)-v*(10**-6))

def cubicsolver(p):
    a=3.11*(10**-7)
    pid=70
    sh=0.04*(2300**0.75)*(0.7**0.33)
    pif=p
    d=1.022*(10**-9)
    k=sh*d/0.026
    K=2.24*(10*5)
    v = Symbol('v')
    df = diff(a*pid*exp(-v*(10**-6)/k)-a*pif*exp(v*(10**-6)*K)-v*(10**-6), v)
    #print(df)
    assumption = 5
    while (True):
        r1 = assumption - (function(assumption,p) / df.subs(v, assumption))
        if (round(assumption,10)==round(r1,10)):
            break

        #print(r1, assumption)
        assumption = r1


    #quo, rem = sympy.div(i * v ** 3 + j * v ** 2 + k * v + l, v - r1, v)

    #a = quo.coeff(v,2)
    #b = quo.coeff(v,1)
    #c = quo.coeff(v,0)

    #r2 = (-b + sympy.sqrt(b ** 2 - 4 * a * c)) / 2 * a
    #r3 = (-b - sympy.sqrt(b ** 2 - 4 * a * c)) / 2 * a

    #rootlist=[r1.__complex__(),r2.__complex__(),r3.__complex__()]
    return r1
x=[]
y=[]

for i in range(10,70):
    x.append(70-i)
    ans=cubicsolver(i)
    y.append(ans)
    print(i,"||",ans)

plt.plot(x,y)
plt.show()

'''def function(v):
    return (v-(3.11*(10**-7))*62.3+(3.11*(10**-7))*2.3*(exp(v/(1.74*(10**-5)))))

def cubicsolver():
    a=0.000000311
    pid=70
    sh=0.04*(2300**0.75)*(0.7**0.33)
    pif=69.5
    d=1.022*(10**-9)
    k=sh*d/0.026
    K=2.24*(10**5)
    v = Symbol('v')
    df = diff(v-(3.11*(10**-7))*62.3+(3.11*(10**-7))*2.3*(exp(v/(1.74*(10**-5)))), v)
    print(df)
    assumption = 0
    while (True):
        r1 = assumption - (function(assumption) / df.subs(v, assumption))
        if (round(assumption,10)==round(r1,10)):
            break

        print(r1, assumption)
        assumption = r1


    #quo, rem = sympy.div(i * v ** 3 + j * v ** 2 + k * v + l, v - r1, v)

    #a = quo.coeff(v,2)
    #b = quo.coeff(v,1)
    #c = quo.coeff(v,0)

    #r2 = (-b + sympy.sqrt(b ** 2 - 4 * a * c)) / 2 * a
    #r3 = (-b - sympy.sqrt(b ** 2 - 4 * a * c)) / 2 * a

    #rootlist=[r1.__complex__(),r2.__complex__(),r3.__complex__()]
    return r1

print(cubicsolver())'''
'''y=[]
x=[]

for i in range(3,70):
    x.append(i)
    y.append(cubicsolver(i,2.3))

plt.plot(x,y)
plt.show()'''